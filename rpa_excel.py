#excel part
import os
import shutil
import openpyxl
import xlrd
import logging
import pandas as pd
def mov_excel():
    source = 'C:/Users/Akshada/Desktop/rpa_idea.xlsx'
    destination = 'C:/Users/Akshada/Desktop/RPA_backend/rpa_idea.xlsx'

    #move from source to destination
    mov = shutil.move(source,destination)
    
def delete_sheet():
    workbook=openpyxl.load_workbook('rpa_idea.xlsx')
    workbook.sheetnames
    str=input('Enter sheet to delete: ')
    std=workbook[str]
    workbook.remove(std)
    workbook.sheetnames
    workbook.save('rpa_idea.xlsx')
    
def create_sheet():
    workbook=openpyxl.load_workbook('rpa_idea.xlsx')
    str=input('Enter sheet name: ')
    workbook.create_sheet(str)
    workbook.sheetnames
    workbook.save('rpa_idea.xlsx')

def rename_sheet():
    workbook=openpyxl.load_workbook('rpa_idea.xlsx')
    workbook.sheetnames
    str=input("Enter sheet name to be renamed: ")
    sheet=workbook[str]
    #type(sheet)
    #sheet.title
    str=input('Enter new sheet name: ')
    sheet.title=str
    workbook.sheetnames
    workbook.save('rpa_idea.xlsx')

def select_sheet():
    try:
        loc = 'C:/Users/Akshada/Desktop/RPA_backend/rpa_idea.xlsx'
        xls=pd.ExcelFile(loc)
        ch=input('Enter sheet to be selected by name: ')
        res = pd.read_excel(xls,ch) 

    #if file is already present
    except xlrd.biffh.XLRDError:
        logger.info("Sheet out of range")

   
while True:
    print(' EXCEL OPTIONS ')
    print('1.MOVE EXCEL FILE')
    print('2.DELETE SHEET IN EXCEL')
    print('3.CREATE SHEET IN EXCEL')
    print('4.RENAME SHEET IN EXCEL')
    print('5.SELECT SHEET IN EXCEL')
    print('6.EXIT')
    ch=int(input('Enter your choice:'))

    logging.basicConfig(filename="logfile.txt",
                    format='%(asctime)s %(message)s',
                    filemode='w')
    #create object
    logger=logging.getLogger()
    logger.setLevel(logging.DEBUG)

    if ch==1:
        mov_excel()
        logger.info('Moved successfully')
    elif ch==2:
        delete_sheet()
        logger.info('Deleted successfully')
    elif ch==3:
        create_sheet()
        logger.info('sheet created successfully')
    elif ch==4:
        rename_sheet()
        logger.info('sheet renamed successfully')
    elif ch==5:
        select_sheet()
        logger.info('sheet selected successfully')
    else:
        break
